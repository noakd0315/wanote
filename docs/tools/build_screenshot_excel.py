# -*- coding: utf-8 -*-
"""Build the screenshot half of the PM's documentation deliverable.

Adds a 「画面一覧」 sheet to wanote_messages.xlsx with one row per screen,
embedding the Japanese and English capture side by side so the two can be
compared without leaving the workbook.

Run:  PYTHONIOENCODING=utf-8 py -3 C:\\Dev\\docs\\tools\\build_screenshot_excel.py
"""
import os

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

EXCEL = r"C:\Dev\docs\wanote_messages.xlsx"
SHOTS = r"C:\Dev\docs\screenshots"
SHEET = "画面一覧"

# slug -> (カテゴリ, 画面名, 備考)
SCREENS = [
    ("01_signin",              "認証", "サインイン画面", "アプリ起動時の最初の画面"),
    ("02_forgot_password",     "認証", "パスワード再設定ダイアログ", "サインイン画面から遷移"),
    ("03_signup",              "認証", "アカウント作成画面", "サインイン画面から切り替え"),
    ("04_language_picker",     "認証", "言語選択ダイアログ", "AppBarの地球アイコンから。端末設定に従う/English/日本語"),
    ("05_signin_filled",       "認証", "サインイン画面（入力済み）", "入力状態の表示確認用"),
    ("06_pet_profile_form",    "認証", "ペットプロフィール登録・編集", "初回登録時は必須。設定からも追加可能"),
    ("10_home",                "アプリシェル", "ホーム画面", "ショートカット5種とボトムナビ"),
    ("11_daily_health",        "日常記録", "健康記録タブ", ""),
    ("12_daily_weight",        "日常記録", "体重タブ", ""),
    ("13_daily_toilet",        "日常記録", "トイレタブ", ""),
    ("14_medical_visits",      "医療", "通院タブ", ""),
    ("15_medical_medications", "医療", "薬タブ", ""),
    ("16_medical_prevention",  "医療", "予防医療タブ", ""),
    ("17_medical_certificates","医療", "証明書タブ", ""),
    ("18_ai_consultation",     "AI", "AI相談タブ", ""),
    ("19_ai_report",           "AI", "レポートタブ", ""),
    ("20_settings",            "設定・課金", "設定画面", "ペット切替/プラン/言語/サインアウト"),
    ("21_language_picker",     "設定・課金", "言語選択ダイアログ（設定内）", "認証画面と同じダイアログ"),
    ("22_pet_switcher",        "設定・課金", "ペット切り替え画面", ""),
    ("24_paywall",             "設定・課金", "有料プラン画面", "サブスク・AIチケット・プロモコード"),
    # --- ダイアログ・入力フォーム ---
    ("30_date_picker",         "ダイアログ", "日付選択", "誕生日・実施日・記録日など全画面共通"),
    ("34_time_picker",         "ダイアログ", "時刻選択", "トイレ記録・投薬リマインダーなど共通"),
    ("31_image_source_sheet",  "ダイアログ", "写真選択シート", "カメラ撮影／フォトライブラリから選択"),
    ("32_weight_add_dialog",   "ダイアログ", "体重の追加", "日常記録→体重のFABから"),
    ("33_urine_dialog",        "ダイアログ", "排尿の記録", "日常記録→トイレ→排尿"),
    ("42_remove_pet_dialog",   "ダイアログ", "ペット削除の確認", "ペット切り替え画面の削除アイコンから"),
    ("36_health_record_form",  "入力フォーム", "健康記録フォーム", "タグ選択・写真添付・コメント"),
    ("35_stool_form",          "入力フォーム", "排便記録フォーム", "硬さ・色・写真・場所"),
    ("37_visit_form",          "入力フォーム", "通院記録フォーム", ""),
    ("38_medication_form",     "入力フォーム", "投薬フォーム", "継続中スイッチ・リマインダー"),
    ("39_prevention_program_form", "入力フォーム", "予防プログラムフォーム", "種別・頻度・有効フラグ"),
    ("41_prevention_record_form",  "入力フォーム", "予防投与記録フォーム", "証明書のAI自動入力導線を含む"),
    ("40_food_portion",        "入力フォーム", "給餌量計算", "ライフステージ・体型・活動レベル"),
    ("43_certificate_list",    "医療", "証明書一覧（登録あり）", "画像が表示されない不具合を修正後"),
    ("44_certificate_viewer",  "ダイアログ", "証明書ビューア", "一覧のカードをタップして拡大表示"),
]

# Captures are 751x1624 (2x DPR); scale down so a row stays readable.
THUMB_H = 420

wb = openpyxl.load_workbook(EXCEL)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET, 1)

headers = ["No.", "カテゴリ", "画面名", "備考", "日本語", "English"]
ws.append(headers)
header_fill = PatternFill("solid", start_color="4A2E1E", end_color="4A2E1E")
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

widths = {"A": 6, "B": 14, "C": 30, "D": 40, "E": 32, "F": 32}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

tmp_dir = os.path.join(SHOTS, "_thumbs")
os.makedirs(tmp_dir, exist_ok=True)

missing = []
row = 2
for i, (slug, category, name, note) in enumerate(SCREENS, start=1):
    ws.cell(row=row, column=1, value=i)
    ws.cell(row=row, column=2, value=category)
    ws.cell(row=row, column=3, value=name)
    ws.cell(row=row, column=4, value=note)
    for c in range(1, 5):
        ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)

    for col, lang in ((5, "ja"), (6, "en")):
        src = os.path.join(SHOTS, f"{slug}_{lang}.png")
        if not os.path.exists(src):
            missing.append(f"{slug}_{lang}")
            ws.cell(row=row, column=col, value="(未取得)")
            continue
        thumb = os.path.join(tmp_dir, f"{slug}_{lang}.png")
        with Image.open(src) as im:
            scale = THUMB_H / im.height
            im.resize((int(im.width * scale), THUMB_H), Image.LANCZOS).save(thumb)
        img = XLImage(thumb)
        ws.add_image(img, f"{get_column_letter(col)}{row}")

    ws.row_dimensions[row].height = THUMB_H * 0.78
    row += 1

ws.freeze_panes = "A2"
wb.save(EXCEL)

print(f"Wrote sheet 「{SHEET}」 with {len(SCREENS)} screens to {EXCEL}")
if missing:
    print("MISSING captures:", ", ".join(missing))
else:
    print("All JA/EN captures embedded.")
