# -*- coding: utf-8 -*-
"""Build an Excel catalog of every localized UI string in wanote.

Reads lib/l10n/app_en.arb + app_ja.arb (source of truth for every
user-facing message in the app) and produces one row per key with:
  key, category (derived from key prefix), description, English, Japanese

This is the reliable/complete half of the "全画面の文書ベース確認"
deliverable -- unlike screenshots, this captures literally every string,
including validation errors, snackbars, and dialogs that only appear in
specific states.
"""
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

REPO = r"C:\Dev\wanote"

with open(rf"{REPO}\lib\l10n\app_en.arb", encoding="utf-8") as f:
    en = json.load(f)
with open(rf"{REPO}\lib\l10n\app_ja.arb", encoding="utf-8") as f:
    ja = json.load(f)

keys = [k for k in en.keys() if not k.startswith("@")]

# Map key prefixes -> human-readable category / screen, ordered so the
# sheet reads roughly in app-navigation order.
CATEGORY_RULES = [
    (r"^(signIn|createAccount|email|password|forgotPassword|referralCode(?!Display|Copied|Applied)|signUp|switchToSign|orDivider|signInWith|cancelButton|sendButton|forcedSignOut|authError)", "認証 / サインイン画面"),
    (r"^biometricGate", "認証 / 生体認証ゲート"),
    (r"^biometricSetup", "認証 / 生体認証セットアップ"),
    (r"^petProfileForm|^petSexOption|^addPetButton", "認証 / ペットプロフィール登録・編集"),
    (r"^(yourPetsScreen|noPetsYet|removePetDialog)", "認証 / ペット切り替え画面"),
    (r"^language", "共通 / 言語切り替え"),
    (r"^(switchPetMenu|upgradePlanMenu|signOutMenu)", "アプリシェル / 設定画面"),
    (r"^nav", "アプリシェル / ボトムナビゲーション"),
    (r"^homeShortcut", "アプリシェル / ホーム画面"),
    (r"^aiSection", "アプリシェル / AIタブ内タブ"),
    (r"^dailyRecord", "アプリシェル / 日常記録タブ内タブ"),
    (r"^imageSource", "共通 / 写真選択シート"),
    (r"^paywall|^restorePurchases|^purchase|^offeringsLoad|^noProductsAvailable|^premium|^aiTickets|^restoreFailedMessage", "課金 / 有料プラン画面"),
    (r"^campaignCode|^referralCodeDisplay|^copyTooltip|^referralCodeCopied|^referralCodeApplied", "課金 / プロモーションコード"),
    (r"^healthRecordTag", "日常記録 / 健康記録タグ"),
    (r"^healthRecordForm|^healthRecordDateTime|^healthRecordTagsSection|^healthRecordPhotos|^healthRecordPhotoCount|^healthRecordComment", "日常記録 / 健康記録フォーム"),
    (r"^healthRecordTimeline|^healthRecordDeleteConfirm|^healthRecordNoComment", "日常記録 / 健康記録一覧・詳細"),
    (r"^toiletFrequencyChart", "日常記録 / トイレ頻度グラフ"),
    (r"^toiletRecordStoolForm|^toiletHardnessSection|^toiletColorSection|^toiletPhoto|^toiletLocation", "日常記録 / 排便記録フォーム"),
    (r"^toiletRecordTimeline|^toiletConsultAi|^toiletUrineColorSubtitle|^toiletStoolConditionSubtitle|^toiletRecordUrineDialog|^toiletUrineColorShade|^toiletRecordSubmitButton|^toiletUrineLabel|^toiletStoolLabel", "日常記録 / トイレ記録一覧・排尿ダイアログ"),
    (r"^toiletHardness|^toiletColor|^urineColor", "日常記録 / 便・尿の状態ラベル"),
    (r"^weight", "日常記録 / 体重記録"),
    (r"^medicalTab|^saveButton$|^requiredFieldValidationError$|^notSetLabel$", "医療 / 共通(タブ・汎用ラベル)"),
    (r"^medicationForm|^medicationName|^medicationDosage|^medicationStartDate|^medicationOngoingSwitch|^medicationEndDate|^medicationReminder", "医療 / 投薬フォーム"),
    (r"^medicationList|^medicationOngoingLabel|^medicationEndDateSubtitle", "医療 / 投薬一覧"),
    (r"^certificateList|^certificateAddPrevention", "医療 / 証明書一覧"),
    (r"^preventionProgramForm|^preventionType|^preventionProductName|^scheduleType|^intervalDays|^numericValue|^preventionProgramActive", "医療 / 予防プログラムフォーム"),
    (r"^preventionProgramList|^scheduleIntervalDays|^preventionProgramInactive", "医療 / 予防プログラム一覧"),
    (r"^savingInProgress|^preventionRecordForm|^administeredAt|^hospitalNameOptional|^nextDueDateLabel|^certificateImageSection|^certificateAlready|^certificateNotRegistered|^certificateCapture|^ocr", "医療 / 予防投与記録フォーム"),
    (r"^vaccinationHistory|^medicationHistory|^vaccinationRecord|^medicationRecord|^preventionRecordList|^nextDueDatePrefix", "医療 / 予防投与記録一覧"),
    (r"^visitForm|^visitedAt|^hospitalNameLabel|^diagnosis|^visitCostLabel|^nextVisitDate", "医療 / 通院記録フォーム"),
    (r"^visitList|^visitFallback|^visitCostYen", "医療 / 通院記録一覧"),
    (r"^aiDisclaimer|^aiEmergency|^aiUpgradeCard", "AI / 共通(免責事項・緊急時)"),
    (r"^consultation", "AI / AI相談画面"),
    (r"^foodPortion|^neuteredStatus|^dogLifeStage|^bodyCondition|^activityLevel", "AI / 給餌量計算画面"),
    (r"^report", "AI / 月次AIレポート画面"),
    (r"^common", "共通 / 汎用ラベル"),
]


def categorize(key: str) -> str:
    for pattern, label in CATEGORY_RULES:
        if re.match(pattern, key):
            return label
    return "未分類"


def description_for(key: str) -> str:
    meta = en.get(f"@{key}")
    if isinstance(meta, dict):
        return meta.get("description", "")
    return ""


rows = []
for key in keys:
    rows.append(
        {
            "category": categorize(key),
            "key": key,
            "description": description_for(key),
            "en": en[key],
            "ja": ja.get(key, ""),
        }
    )

# Sort by category (in CATEGORY_RULES order), then by key for stable,
# readable grouping.
category_order = {label: i for i, (_, label) in enumerate(CATEGORY_RULES)}
rows.sort(key=lambda r: (category_order.get(r["category"], 999), r["key"]))

wb = Workbook()
ws = wb.active
ws.title = "メッセージ一覧"

headers = ["No.", "カテゴリ / 画面", "キー", "説明", "英語 (English)", "日本語"]
ws.append(headers)
header_fill = PatternFill(start_color="4A2E1E", end_color="4A2E1E", fill_type="solid")
for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)

prev_category = None
for i, r in enumerate(rows, start=1):
    row_idx = i + 1
    ws.cell(row=row_idx, column=1, value=i)
    ws.cell(row=row_idx, column=2, value=r["category"])
    ws.cell(row=row_idx, column=3, value=r["key"])
    ws.cell(row=row_idx, column=4, value=r["description"])
    ws.cell(row=row_idx, column=5, value=r["en"])
    ws.cell(row=row_idx, column=6, value=r["ja"])
    for col in range(1, 7):
        ws.cell(row=row_idx, column=col).alignment = Alignment(
            vertical="top", wrap_text=True
        )
    # Light shading band per category for readability.
    if r["category"] != prev_category:
        band_fill = PatternFill(
            start_color="F3E9DE", end_color="F3E9DE", fill_type="solid"
        )
        for col in range(1, 7):
            existing = ws.cell(row=row_idx, column=col)
            existing.fill = band_fill
    prev_category = r["category"]

widths = {"A": 6, "B": 28, "C": 34, "D": 46, "E": 46, "F": 40}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{len(rows)+1}"

# Summary sheet
summary = wb.create_sheet("サマリー", 0)
summary["A1"] = "wanote 全画面メッセージ一覧"
summary["A1"].font = Font(bold=True, size=16)
summary["A3"] = "総メッセージ数"
summary["B3"] = len(rows)
summary["A4"] = "カテゴリ数"
summary["B4"] = len({r["category"] for r in rows})
summary["A5"] = "生成元"
summary["B5"] = "lib/l10n/app_en.arb, lib/l10n/app_ja.arb"
summary["A7"] = "カテゴリ別内訳"
summary["A7"].font = Font(bold=True)
cat_counts = {}
for r in rows:
    cat_counts[r["category"]] = cat_counts.get(r["category"], 0) + 1
row = 8
for label in sorted(cat_counts, key=lambda l: category_order.get(l, 999)):
    summary.cell(row=row, column=1, value=label)
    summary.cell(row=row, column=2, value=cat_counts[label])
    row += 1
summary.column_dimensions["A"].width = 34
summary.column_dimensions["B"].width = 14

out_path = r"C:\Users\tomoc\AppData\Local\Temp\claude\C--Dev\74f3a2ed-83b2-473a-86f2-15c93dfaca15\scratchpad\wanote_messages.xlsx"
wb.save(out_path)
print(f"Saved {len(rows)} rows to {out_path}")
