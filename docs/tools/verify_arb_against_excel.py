# -*- coding: utf-8 -*-
"""Verify that every PM instruction in the Excel's G column ("修正後") has
actually been applied to lib/l10n/app_ja.arb.

Guards against the failure mode that already happened once: the Excel was
updated and reported as done while the app itself still showed the old text.

Usage (from anywhere):
    PYTHONIOENCODING=utf-8 py -3 C:\\Dev\\docs\\tools\\verify_arb_against_excel.py

Exits non-zero if anything is unapplied, so it can gate a commit.
"""
import json
import sys

import openpyxl

EXCEL = r"C:\Dev\docs\wanote_messages.xlsx"
ARB = r"C:\Dev\wanote\lib\l10n\app_ja.arb"
SHEET = "メッセージ一覧"
KEY_COL, FIX_COL = 3, 7
# G-column values starting with this mean "reviewed, deliberately unchanged".
NO_CHANGE_PREFIX = "対応不要"

ws = openpyxl.load_workbook(EXCEL)[SHEET]
ja = json.load(open(ARB, encoding="utf-8"))

applied, skipped, pending = [], [], []
for row in range(2, ws.max_row + 1):
    key = ws.cell(row=row, column=KEY_COL).value
    fix = ws.cell(row=row, column=FIX_COL).value
    if not key or not fix:
        continue
    if str(fix).startswith(NO_CHANGE_PREFIX):
        skipped.append((row, key))
    elif ja.get(key) == fix:
        applied.append((row, key))
    else:
        pending.append((row, key, ja.get(key), fix))

total = len(applied) + len(skipped) + len(pending)
print(f"Excel G列の修正指示: {total} 件")
print(f"  ARB反映済み : {len(applied)}")
print(f"  対応不要    : {len(skipped)}")
print(f"  未反映      : {len(pending)}")

for row, key, current, expected in pending:
    print(f"\n  [未反映] 行{row} {key}")
    print(f"    現在  : {current!r}")
    print(f"    あるべき: {expected!r}")

if pending:
    print("\nNG: 未反映の修正があります。fix_arb.py 系で反映してください。")
    sys.exit(1)
print("\nOK: すべての修正が反映されています。")
